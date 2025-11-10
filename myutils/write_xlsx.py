# -*- coding: utf-8 -*-
'''
@File : write_xlsx.py
@Time : 2025/06/04 10:54:13
@Author : wlj 
'''

import os
import time
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
# from openpyxl.chart import PieChart, Reference, ScatterChart, Series
# from openpyxl.chart.label import DataLabelList


def set_style(stylename, fontName='Microsoft YaHei',bold=False,italic=False, horizontal='center', 
              color='000000', fill=False, fillcolor="B8F0C1", front_size=10, border=False):# little green by fault
    my_style = NamedStyle(name=stylename)
    # 'Times New Roman', '黑体'
    font = Font(name=fontName, italic=italic, bold=bold, color=color, size=front_size)  # color='000000'=black, size=12
    # 设置水平居中, 设置垂直居中, 设置后遇到换行符将自动换行
    alignment = Alignment(horizontal=horizontal, vertical='center', wrapText=True)
    # wrapText=True:设置自动换行
    my_style.font = font
    my_style.alignment = alignment

    if fill:
        my_style.fill = PatternFill(fill_type="solid", fgColor=fillcolor) 
    if border:
        side = Side(style='thin', color='000000') # 'thin','thick', 'dashed', '000000':black
        my_style.border = Border(left=side, right=side, top=side, bottom=side)
    return my_style


def write_xlsx_preface(mysheet, style, start_row=1, max_column=9):
    data = "文本数据量"
    key_count = "关键词数量"
    key_merge = "关键词是否合并"
    line_process = "文本处理方式\n" \
                   "0: 整体处理   1: 逐行处理"
    match_method = "匹配条件"
    match_case = "匹配大小写"
    match_wholeword = "匹配全词"
    time_elapse = "耗时(S)"
    # memory_elapse = "内存消耗"
    extral = "备注"
    test_time = "测试时间: "

    test_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_column) # 
    mysheet.cell(start_row, 1, value=f"{test_time} : {test_time_str}").style = style   

    start_row += 1
    end_row = start_row + 1  
    
    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1) # 
    mysheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2) # 
    mysheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3) # 
    mysheet.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4) # 
    mysheet.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=6) # 
    mysheet.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7) # 
    mysheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8) # 
    mysheet.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9) # 
    

    mysheet.cell(start_row, 1, value=data).style = style      
    mysheet.cell(start_row, 2, value=key_count).style = style
    mysheet.cell(start_row, 3, value=key_merge).style = style      
    mysheet.cell(start_row, 4, value=line_process).style = style
    mysheet.cell(start_row, 5, value=match_method).style = style   
    mysheet.cell(start_row + 1, 5, value=match_case).style = style   
    mysheet.cell(start_row + 1, 6, value=match_wholeword).style = style      
    mysheet.cell(start_row, 7, value=time_elapse).style = style     
    # mysheet.cell(start_row, 8, value=memory_elapse).style = style
    mysheet.cell(start_row, 8, value=extral).style = style
    
    return mysheet, end_row

'''
******************************************************************
args:
    configs = {'match_case': match_case,  # 匹配大小写
               'match_wholeword': match_wholeword,  # 全词匹配
               'line_by_line': line_by_line,  # 逐行处理
               'keyword_merge': keyword_merge,  # 关键词是否合并
              }
    keywords:  a list of key words
return:
    
******************************************************************
'''
def write_content(mysheet, style, configs, key_words, time_elaspe, row=1):
    
    mysheet.cell(row, 2, value=len(key_words)).style = style
    mysheet.cell(row, 3, value=configs['keyword_merge']).style = style
    mysheet.cell(row, 4, value=configs['line_by_line']).style = style
    mysheet.cell(row, 5, value=configs['match_case']).style = style
    mysheet.cell(row, 6, value=configs['match_wholeword']).style = style
    mysheet.cell(row, 7, value="%.2f"%time_elaspe).style = style

    return mysheet
  

def write_xlsx_eval_preface(mysheet, style, start_row=1, max_column=8):
    model = "Model"
    tasks = "Tasks"
    metric = "metirc"
    platform = ["transformers", "vllm(torch.bfloat16)", "vllm(FP8)", "vllm(AWQ:W4A16)"]
    acc = "acc"
    acc_norm = "acc_norm"
    extral = "备注"
    test_time = "测试时间: "
    time_ = "runtime"

    test_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_column) # 
    mysheet.cell(start_row, 1, value=f"{test_time} : {test_time_str}").style = style   

    start_row += 1
    end_row = start_row + 2

    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1) # 
    mysheet.cell(start_row, 1, value=tasks).style = style 
    mysheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2) # 
    mysheet.cell(start_row, 2, value=model).style = style 
    mysheet.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=max_column) # metric
    mysheet.cell(start_row, 3, value=metric).style = style 

    for i in range(len(platform)):
        start_col = 3 + i * 3
        mysheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2) # 
        mysheet.cell(start_row+1, start_col, value=platform[i]).style = style 

        mysheet.cell(start_row+2, start_col, value=acc).style = style 
        mysheet.cell(start_row+2, start_col+1, value=acc_norm).style = style
        mysheet.cell(start_row+2, start_col+2, value=time_).style = style
    
    return mysheet, end_row

def write_llm_content(mysheet, style, results, row=1, col=1):
    for i, (datasets,r) in enumerate(results['results'].items()):
        print('datasets = %s, results = %s' %(datasets, r))
        
        model = results['configs'][datasets]['metadata']['pretrained']
        print("write_llm_content, model = %s" % model)
        model = os.path.basename(model)

        result_list = list(r.values())
    
        time_ = r['runtime']
        acc = "%.4f ± %.4f" % (result_list[0], result_list[1])
        acc_norm = "%.4f ± %.4f" % (result_list[2], result_list[3])
        
        # if datasets == 'hellaswag':
        #     acc = "%.4f ± %.4f" % (r['acc,none'], r['acc_stderr,none'])
        #     acc_norm = "%.4f ± %.4f" % (r['acc_norm,none'], r['acc_norm_stderr,none'])
        # elif datasets == 'drop':
        #     acc = "%.4f ± %.4f" % (r['em,none'], r['em_stderr,none'])
        #     acc_norm = "%.4f ± %.4f" % (r['f1,none'], r['f1_norm_stderr,none'])
        
        mysheet.cell(row + i, 1, value=datasets).style = style
        mysheet.cell(row + i, 2, value=model).style = style
        mysheet.cell(row + i, col, value=acc).style = style
        mysheet.cell(row + i, col + 1, value=acc_norm).style = style
        mysheet.cell(row + i, col + 2, value=time_).style = style

        

    return mysheet




if __name__ == '__main__':
    a = 0
    

